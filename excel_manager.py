"""讀寫名片 Excel 資料庫。"""

from __future__ import annotations

import re
from datetime import date, datetime
from pathlib import Path
from urllib.parse import unquote, urlparse

import openpyxl
from openpyxl.styles import Font
from openpyxl.worksheet.worksheet import Worksheet

SHEET_NAME = "名片資料庫"
SCAN_CARD_HEADER = "Scan_card"
LEGACY_PHOTO_HEADER = "名片照片"
UPLOAD_FOLDER_NAME = "file Name_Card_system"


def scan_card_link_text(card_id: str) -> str:
    return f"Scan_card_{card_id}"

HEADERS = [
    "卡片ID",
    "Name",
    "Company_name",
    "Title",
    "Tel",
    "FEX",
    "company_Tel",
    "Email",
    "Company Address",
    "Company Website",
    "Line/ID",
    "Wechat_ID",
    "備註",
    "Scan_date",
    "Event",
    SCAN_CARD_HEADER,
]


class ExcelManager:
    def __init__(self, excel_path: str | Path) -> None:
        self.excel_path = Path(excel_path).resolve()
        self.workbook = openpyxl.load_workbook(self.excel_path)
        self.sheet: Worksheet = self.workbook[SHEET_NAME]
        if self._migrate_scan_card_column():
            self.save()
        self._header_index = self._build_header_index()

    def _build_header_index(self) -> dict[str, int]:
        mapping: dict[str, int] = {}
        for col in range(1, self.sheet.max_column + 1):
            header = self.sheet.cell(row=1, column=col).value
            if header:
                mapping[str(header)] = col
        return mapping

    def _scan_card_column(self) -> int:
        if SCAN_CARD_HEADER in self._header_index:
            return self._header_index[SCAN_CARD_HEADER]
        raise KeyError(f"找不到欄位：{SCAN_CARD_HEADER}")

    @staticmethod
    def _photo_name_from_target(target: str) -> str:
        if target.startswith("file:"):
            path = unquote(urlparse(target).path)
            if (
                len(path) > 2
                and path[0] == "/"
                and path[2] == ":"
            ):
                path = path[1:]
            return Path(path).name
        return Path(unquote(target)).name

    def _photo_name_from_cell(self, cell) -> str | None:
        if cell.hyperlink and cell.hyperlink.target:
            return self._photo_name_from_target(cell.hyperlink.target)
        value = str(cell.value).strip() if cell.value else ""
        if not value or value.startswith("Scan_card"):
            return None
        return value

    def _resolve_scan_card_file(
        self,
        photo_name: str,
        photo_path: Path | None = None,
    ) -> Path | None:
        base_dir = self.excel_path.parent
        candidates: list[Path] = []
        if photo_path is not None:
            candidates.append(Path(photo_path))
        candidates.append(base_dir / UPLOAD_FOLDER_NAME / photo_name)
        candidates.append(base_dir / photo_name)

        seen: set[Path] = set()
        for candidate in candidates:
            resolved = candidate.resolve()
            if resolved in seen:
                continue
            seen.add(resolved)
            if resolved.is_file():
                return resolved
        return None

    def _apply_scan_card_hyperlink(self, cell, file_path: Path, card_id: str) -> None:
        cell.value = scan_card_link_text(card_id)
        cell.hyperlink = file_path.resolve().as_uri()
        cell.font = Font(color="0563C1", underline="single")

    def _migrate_scan_card_column(self) -> bool:
        headers = {
            self.sheet.cell(row=1, column=col).value: col
            for col in range(1, self.sheet.max_column + 1)
            if self.sheet.cell(row=1, column=col).value
        }

        if SCAN_CARD_HEADER in headers:
            scan_col = headers[SCAN_CARD_HEADER]
        elif LEGACY_PHOTO_HEADER in headers:
            scan_col = headers[LEGACY_PHOTO_HEADER]
            self.sheet.cell(row=1, column=scan_col).value = SCAN_CARD_HEADER
        else:
            scan_col = self.sheet.max_column + 1
            self.sheet.cell(row=1, column=scan_col).value = SCAN_CARD_HEADER

        changed = scan_col == self.sheet.max_column and (
            self.sheet.cell(row=1, column=scan_col).value == SCAN_CARD_HEADER
        )
        if LEGACY_PHOTO_HEADER in headers:
            changed = True

        id_col = headers.get("卡片ID")

        for row in range(2, self.sheet.max_row + 1):
            cell = self.sheet.cell(row=row, column=scan_col)
            photo_name = self._photo_name_from_cell(cell)
            if not photo_name:
                continue
            resolved = self._resolve_scan_card_file(photo_name)
            if not resolved:
                continue
            card_id = ""
            if id_col:
                card_id = str(self.sheet.cell(row=row, column=id_col).value or "").strip()
            if not card_id:
                continue
            target_uri = resolved.as_uri()
            link_text = scan_card_link_text(card_id)
            current_target = cell.hyperlink.target if cell.hyperlink else None
            if cell.value != link_text or current_target != target_uri:
                self._apply_scan_card_hyperlink(cell, resolved, card_id)
                changed = True

        return changed

    def get_existing_photo_names(self) -> set[str]:
        col = self._scan_card_column()
        names: set[str] = set()
        for row in range(2, self.sheet.max_row + 1):
            cell = self.sheet.cell(row=row, column=col)
            photo_name = self._photo_name_from_cell(cell)
            if photo_name:
                names.add(photo_name)
        return names

    def get_next_card_id(self) -> str:
        col = self._header_index["卡片ID"]
        max_num = 0
        for row in range(2, self.sheet.max_row + 1):
            value = self.sheet.cell(row=row, column=col).value
            if not value:
                continue
            match = re.match(r"BC(\d+)", str(value).strip(), re.I)
            if not match:
                continue
            max_num = max(max_num, int(match.group(1)))
        return f"BC{max_num + 1:03d}"

    def find_empty_row(self) -> tuple[int, str | None]:
        id_col = self._header_index["卡片ID"]
        name_col = self._header_index["Name"]
        scan_col = self._scan_card_column()
        for row in range(2, self.sheet.max_row + 1):
            card_id = self.sheet.cell(row=row, column=id_col).value
            name = self.sheet.cell(row=row, column=name_col).value
            scan_cell = self.sheet.cell(row=row, column=scan_col)
            has_scan = bool(self._photo_name_from_cell(scan_cell))
            if name or has_scan:
                continue
            if card_id:
                return row, str(card_id).strip()
            return row, None
        return self.sheet.max_row + 1, None

    def add_card(
        self,
        card_id: str,
        row_data: dict[str, str | None],
        scan_date: str | date | datetime,
        event: str,
        photo_name: str,
        photo_path: str | Path | None = None,
    ) -> int:
        row_idx, existing_id = self.find_empty_row()
        card_id = existing_id or card_id
        values = {
            "卡片ID": card_id,
            "Scan_date": self._format_scan_date(scan_date),
            "Event": event,
            **row_data,
        }
        for header, value in values.items():
            if header not in self._header_index:
                continue
            self.sheet.cell(
                row=row_idx,
                column=self._header_index[header],
                value=value,
            )

        resolved = self._resolve_scan_card_file(
            photo_name,
            Path(photo_path) if photo_path else None,
        )
        if not resolved:
            raise FileNotFoundError(
                f"找不到名片 PDF：{photo_name}（請確認檔案在 {UPLOAD_FOLDER_NAME} 資料夾）"
            )

        scan_cell = self.sheet.cell(row=row_idx, column=self._scan_card_column())
        self._apply_scan_card_hyperlink(scan_cell, resolved, card_id)
        return row_idx

    def save(self) -> None:
        self.workbook.save(self.excel_path)

    @staticmethod
    def _format_scan_date(scan_date: str | date | datetime) -> str | date | datetime:
        if isinstance(scan_date, (date, datetime)):
            return scan_date
        text = scan_date.strip()
        for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%Y/%m/%d"):
            try:
                return datetime.strptime(text, fmt).date()
            except ValueError:
                continue
        return text
